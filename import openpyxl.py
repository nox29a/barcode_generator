import openpyxl
from openpyxl import load_workbook
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from barcode import Code128
from barcode.writer import ImageWriter
from PIL import Image, ImageFont
import re
import os

# Wczytanie pliku Excela
excel_file = 'nazwa_pliku.xlsx'  # Podaj ścieżkę do swojego pliku Excela
sheet_name = 'Arkusz1'  # Nazwa arkusza, na którym znajduje się kolumna do odczytu

wb = load_workbook(excel_file, data_only=True)
sheet = wb[sheet_name]

# Kolumna do odczytu (zmień na odpowiednią kolumnę)
column_index = 'B'

# Nazwa pliku PDF
pdf_file = 'kody_kreskowe.pdf'

# Rozpoczęcie tworzenia pliku PDF
c = canvas.Canvas(pdf_file, pagesize=letter)
pdf_data = []

# Font do zapisywania jako obrazy
font = ImageFont.load_default()

# Funkcja do usuwania niedozwolonych znaków z nazwy pliku
def sanitize_filename(filename):
    return re.sub(r'[<>:"/\\|?*]', '_', filename)

for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=sheet[column_index + '2'].column, max_col=sheet[column_index + str(sheet.max_row)].column):
    cell_value = str(row[0].value)
    
    # Generowanie kodu kreskowego Code128
    code128 = Code128(cell_value, writer=ImageWriter())
    barcode_image = code128.render()
    
    # Usuwanie niedozwolonych znaków z nazwy pliku
    sanitized_filename = sanitize_filename(cell_value)
    
    # Zapisanie kodu kreskowego jako obrazu PNG
    png_file = f'barcode_{sanitized_filename}.png'
    barcode_image.save(png_file)
    
    pdf_data.append(png_file)

# Dodanie kodów kreskowych do strony PDF
for png_file in pdf_data:
    c.drawImage(png_file, 100, 600, width=400, height=100)
    
    c.showPage()

c.save()
# Usunięcie plików PNG po ich umieszczeniu w pliku PDF
for png_file in pdf_data:
    os.remove(png_file)

# Zapisanie danych do pliku PDF

print(f"Kody kreskowe zostały zapisane jako strony w pliku PDF: {pdf_file}")
